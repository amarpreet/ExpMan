from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
import os
import pandas as pd
import csv
import openpyxl
import pyexcel as pe
from tempfile import NamedTemporaryFile
from pydantic import BaseModel
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import re

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
OUTPUT_DIR = "outputs"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

@app.post("/upload/")
def upload_file(
    bank_file: UploadFile = File(...),
    hist_file: UploadFile = File(None),
    account_name: str = Form(...),
    reconciled: str = Form("Y")
):
    # Save uploaded files
    bank_path = os.path.join(UPLOAD_DIR, bank_file.filename)
    with open(bank_path, "wb") as f:
        f.write(bank_file.file.read())
    hist_path = None
    if hist_file:
        hist_path = os.path.join(UPLOAD_DIR, hist_file.filename)
        with open(hist_path, "wb") as f:
            f.write(hist_file.file.read())

    # --- Step 1: Parse the uploaded bank file and detect columns ---
    try:
        def find_header_row(data, max_scan=10):
            # Find the first row with mostly non-empty, unique values
            for i, row in enumerate(data[:max_scan]):
                non_empty = [cell for cell in row if str(cell).strip() != '']
                if len(non_empty) >= max(3, len(row)//2) and len(set(non_empty)) == len(non_empty):
                    return i
            return 0  # fallback to first row

        if bank_file.filename.lower().endswith(".csv"):
            with open(bank_path, newline='', encoding='utf-8') as f:
                reader = list(csv.reader(f))
            header_row = find_header_row(reader)
            df = pd.DataFrame(reader[header_row+1:], columns=reader[header_row])
        elif bank_file.filename.lower().endswith(".xlsx"):
            wb = openpyxl.load_workbook(bank_path, read_only=True)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            header_row = find_header_row(data)
            df = pd.DataFrame(data[header_row+1:], columns=data[header_row])
        elif bank_file.filename.lower().endswith(".xls"):
            sheet = pe.get_sheet(file_name=bank_path)
            data = sheet.to_array()
            header_row = find_header_row(data)
            df = pd.DataFrame(data[header_row+1:], columns=data[header_row])
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload a CSV, XLS, or XLSX file.")
    except Exception as e:
        print(f"Error reading file: {e}")  # Print error to server log
        raise HTTPException(status_code=400, detail=f"Error reading file: {str(e)}")

    # Get column names and a preview (first 5 rows)
    columns = list(df.columns)
    preview = df.head(5).to_dict(orient="records")

    return {
        "bank_file": bank_file.filename,
        "hist_file": hist_file.filename if hist_file else None,
        "account_name": account_name,
        "reconciled": reconciled,
        "columns": columns,
        "preview": preview
    }

class OutputRequest(BaseModel):
    bank_file: str
    account_name: str
    reconciled: str
    date_col: str
    desc_col: str
    amt_col: str
    hist_file: str = None  # Optional historical file

@app.post("/generate_output/")
def generate_output(req: OutputRequest):
    print("/generate_output/ endpoint called")
    bank_path = os.path.join(UPLOAD_DIR, req.bank_file)
    # Re-read the file using the same logic as before
    try:
        def find_header_row(data, max_scan=10):
            # Find the first row with mostly non-empty, unique values
            for i, row in enumerate(data[:max_scan]):
                non_empty = [cell for cell in row if str(cell).strip() != '']
                if len(non_empty) >= max(3, len(row)//2) and len(set(non_empty)) == len(non_empty):
                    return i
            return 0  # fallback to first row

        if req.bank_file.lower().endswith(".csv"):
            with open(bank_path, newline='', encoding='utf-8') as f:
                reader = list(csv.reader(f))
            header_row = find_header_row(reader)
            df = pd.DataFrame(reader[header_row+1:], columns=reader[header_row])
        elif req.bank_file.lower().endswith(".xlsx"):
            wb = openpyxl.load_workbook(bank_path, read_only=True)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            header_row = find_header_row(data)
            df = pd.DataFrame(data[header_row+1:], columns=data[header_row])
        elif req.bank_file.lower().endswith(".xls"):
            sheet = pe.get_sheet(file_name=bank_path)
            data = sheet.to_array()
            header_row = find_header_row(data)
            df = pd.DataFrame(data[header_row+1:], columns=data[header_row])
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload a CSV, XLS, or XLSX file.")
    except Exception as e:
        print(f"Error reading file: {e}")
        raise HTTPException(status_code=400, detail=f"Error reading file: {str(e)}")

    # --- Step 2: Parse historical categorization file if provided ---
    hist_df = None
    if req.hist_file:
        hist_path = os.path.join(UPLOAD_DIR, req.hist_file)
        try:
            if req.hist_file.lower().endswith(".csv"):
                hist_df = pd.read_csv(
                    hist_path,
                    header=0,  # Use the first row as header, skip it as data
                    dtype=str,
                    keep_default_na=False
                )
            elif req.hist_file.lower().endswith(".xlsx"):
                hist_df = pd.read_excel(
                    hist_path,
                    header=0,  # Use the first row as header
                    dtype=str
                )
            elif req.hist_file.lower().endswith(".xls"):
                hist_df = pd.read_excel(
                    hist_path,
                    header=0,  # Use the first row as header
                    dtype=str
                )
            # Remove rows where Notes is empty or is the string 'Notes'
            hist_df = hist_df[hist_df["Notes"].str.strip().str.lower() != "notes"]
            hist_df = hist_df[hist_df["Notes"].str.strip() != ""]
            print(f"Loaded hist_df shape: {hist_df.shape}")
            print(hist_df.head())

            # Dynamically detect column names for Details, Category, Notes
            def find_col(possible_names, columns):
                for name in possible_names:
                    for col in columns:
                        if col.strip().lower() == name.strip().lower():
                            return col
                return None
            details_col = find_col(["Details", "Details / [To]"], hist_df.columns)
            category_col = find_col(["Category"], hist_df.columns)
            notes_col = find_col(["Notes"], hist_df.columns)
        except Exception as e:
            print(f"Error reading historical file: {e}")
            hist_df = None
            details_col = category_col = notes_col = None

    # Prepare output DataFrame
    output = pd.DataFrame()
    output["Account Name"] = [req.account_name] * len(df)
    output["Date"] = pd.to_datetime(df[req.date_col], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y')

    # --- Step 3: Suggest Details and Category based on historical data ---
    def jaccard_similarity(a, b):
        set_a = set(re.findall(r'\w+', a.lower()))
        set_b = set(re.findall(r'\w+', b.lower()))
        if not set_a or not set_b:
            return 0.0
        return len(set_a & set_b) / len(set_a | set_b)

    def suggest_details_and_category(note):
        print(f"suggest_details_and_category called for note: {note}")
        if hist_df is not None and not hist_df.empty and notes_col and details_col and category_col:
            note_clean = str(note).strip().lower()
            hist_notes = hist_df[notes_col].astype(str).fillna("").tolist()
            best_score = 0.0
            best_idx = -1
            for idx, hist_note in enumerate(hist_notes):
                hist_note_clean = hist_note.strip().lower()
                # Substring match (strongest)
                if note_clean in hist_note_clean or hist_note_clean in note_clean:
                    score = 1.0
                else:
                    score = jaccard_similarity(note_clean, hist_note_clean)
                print(f"  {idx}: {hist_note} | Score: {score}")
                if score > best_score:
                    best_score = score
                    best_idx = idx
            print(f"Best score for '{note}': {best_score} (idx: {best_idx})")
            if best_score > 0.3 and best_idx != -1:
                details = hist_df.iloc[best_idx][details_col] if pd.notnull(hist_df.iloc[best_idx][details_col]) else ""
                category = hist_df.iloc[best_idx][category_col] if pd.notnull(hist_df.iloc[best_idx][category_col]) else ""
                return details, category
            else:
                print("No good match found.")
        else:
            print("No historical DataFrame available or it is empty, or columns not found.")
        return "", ""

    details_list = []
    category_list = []
    for note in df[req.desc_col]:
        details, category = suggest_details_and_category(note)
        details_list.append(details)
        category_list.append(category)
    output["Details"] = details_list
    output["Category"] = category_list
    output["Notes"] = df[req.desc_col]
    output["Cheque/Check Number"] = ""
    output["Amount"] = pd.to_numeric(df[req.amt_col], errors='coerce')
    output["Reconciled"] = [req.reconciled] * len(df)
    output = output.sort_values(by="Date")
    output_file = f"output_{req.account_name}.csv"
    output_file_path = os.path.join(OUTPUT_DIR, output_file)
    output.to_csv(output_file_path, index=False, header=False)
    return {"output_file": output_file}

@app.get("/download/{filename}")
def download_file(filename: str):
    file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found.")
    return FileResponse(file_path, media_type='text/csv', filename=filename)
